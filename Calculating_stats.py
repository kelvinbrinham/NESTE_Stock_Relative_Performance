import pandas as pd
import numpy as np
import openpyxl as xl
import datetime
from datetime import datetime as dt
import xlsxwriter as xlw
import matplotlib.pyplot as plt
import dateutil.relativedelta as reldel
import statistics as stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import functions as func
import os

#Set how much of a dataframe is displayed when running the scipt
pd.set_option('display.max_rows', 10)
pd.set_option('display.max_columns', 10)

#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>

from Calculating_outperformance import Neste_and_SXXP_price_df

#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>

#MAKE YES COLUMNS YELLOW!
Complete_data_coloured_file_name = 'OUTPUT/Complete_data_coloured.xlsx'
Neste_and_SXXP_price_df.to_excel(Complete_data_coloured_file_name, sheet_name='Sheet1', startrow = 2, index=False)
func.colour_yes_cells_yellow(Complete_data_coloured_file_name)

#Formatting Column titles in Complete_data_coloured.xlsx
Complete_data_coloured_wb = xl.load_workbook(Complete_data_coloured_file_name)
Complete_data_coloured_ws = Complete_data_coloured_wb.active

Complete_data_coloured_ws['B3'] = 'NESTE FH last PX'
Complete_data_coloured_ws['C3'] = 'SXXP last PX'
Complete_data_coloured_ws['E3'] = 'RP 3-month High'
Complete_data_coloured_ws['F3'] = 'RP 3-month High Date'
Complete_data_coloured_ws['G3'] = 'RP 3-month High (y/n)'
Complete_data_coloured_ws['H3'] = '1-month Performance'
Complete_data_coloured_ws['I3'] = '3-month Performance'
Complete_data_coloured_ws['J3'] = '6-month Performance'

Complete_data_coloured_wb.save(Complete_data_coloured_file_name)

#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>
#Create new data frame only containing 3 month highs
#I.e. data frame with only the 'yes' rows

#Drop non Yes rows
Neste_and_SXXP_HIGH_DATA_df = Neste_and_SXXP_price_df.drop(Neste_and_SXXP_price_df[Neste_and_SXXP_price_df.RP_3_month_high_y_or_n != 'Yes'].index)

#Drop second date column and yes column
Neste_and_SXXP_HIGH_DATA_df.drop(['RP', 'RP_3_month_high_date', 'RP_3_month_high_y_or_n'], axis=1, inplace=True)

#Minus 1 from percentage performances
one_month_outperformance_lst = list(Neste_and_SXXP_HIGH_DATA_df['1_month_outperformance'])
three_month_outperformance_lst = list(Neste_and_SXXP_HIGH_DATA_df['3_month_outperformance'])
six_month_outperformance_lst = list(Neste_and_SXXP_HIGH_DATA_df['6_month_outperformance'])

one_month_outperformance_lst = [x-1 for x in one_month_outperformance_lst]
three_month_outperformance_lst = [x-1 for x in three_month_outperformance_lst]
six_month_outperformance_lst = [x-1 for x in six_month_outperformance_lst if not np.isnan(x)] + ([np.nan] * 12)

Neste_and_SXXP_HIGH_DATA_df['1_month_outperformance'] = one_month_outperformance_lst
Neste_and_SXXP_HIGH_DATA_df['3_month_outperformance'] = three_month_outperformance_lst
Neste_and_SXXP_HIGH_DATA_df['6_month_outperformance'] = six_month_outperformance_lst

Neste_and_SXXP_HIGH_DATA_df.to_excel('OUTPUT/Three_month_high_data.xlsx', sheet_name='Sheet1', startrow = 2, index=False)


Three_month_high_data_wb_file_name = 'OUTPUT/Three_month_high_data_formatted.xlsx'


Neste_and_SXXP_HIGH_DATA_df.to_excel(Three_month_high_data_wb_file_name, sheet_name='Sheet1', startrow = 2, index=False)

Three_month_high_data_wb = xl.load_workbook(Three_month_high_data_wb_file_name)
Three_month_high_data_ws = Three_month_high_data_wb.active

fill_cell_green = PatternFill(patternType='solid',
                           fgColor='35FC03')

fill_cell_red = PatternFill(patternType='solid',
                           fgColor='FC2C03')


#Format percentage cells accordingly and colour
percentage_column_letters_lst = ['E', 'F', 'G']
for letter in percentage_column_letters_lst:
    for i in range(4, 263 + 1, 1): #4-263
        if type(Three_month_high_data_ws[letter + str(i)].value) == float:
            if Three_month_high_data_ws[letter + str(i)].value >= 0:
                Three_month_high_data_ws[letter + str(i)].fill = fill_cell_green
            else:
                Three_month_high_data_ws[letter + str(i)].fill = fill_cell_red


        Three_month_high_data_ws[letter + str(i)].number_format = '0.00%'

#Formatting sheet headers for readability
Three_month_high_data_ws['B3'] = 'NESTE FH last PX'
Three_month_high_data_ws['C3'] = 'SXXP last PX'
Three_month_high_data_ws['D3'] = 'RP 3-month High'
Three_month_high_data_ws['E3'] = '1-month Performance'
Three_month_high_data_ws['F3'] = '3-month Performance'
Three_month_high_data_ws['G3'] = '6-month Performance'

Three_month_high_data_wb.save(Three_month_high_data_wb_file_name)

os.remove('OUTPUT/Three_month_high_data.xlsx')

#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>
one_month_observations = func.length_ignoring_strings_and_nan(one_month_outperformance_lst)
three_month_observations = func.length_ignoring_strings_and_nan(three_month_outperformance_lst)
six_month_observations = func.length_ignoring_strings_and_nan(six_month_outperformance_lst)
observations_lst = [one_month_observations, three_month_observations, six_month_observations]


avg_1_month_out_performance = func.list_mean_ignoring_strings_and_nan(one_month_outperformance_lst)
avg_3_month_out_performance = func.list_mean_ignoring_strings_and_nan(three_month_outperformance_lst)
avg_6_month_out_performance = func.list_mean_ignoring_strings_and_nan(six_month_outperformance_lst)
avg_out_performance_lst = [avg_1_month_out_performance, avg_3_month_out_performance, avg_6_month_out_performance]

median_1_month_out_performance = func.list_median_ignoring_strings_and_nan(one_month_outperformance_lst)
median_3_month_out_performance = func.list_median_ignoring_strings_and_nan(three_month_outperformance_lst)
median_6_month_out_performance = func.list_median_ignoring_strings_and_nan(six_month_outperformance_lst)
median_out_performance_lst = [median_1_month_out_performance, median_3_month_out_performance, median_6_month_out_performance]

#Calculate Hit ratio = no_outperformances / no_underperformances
#I define outperformance as RP >= 0 and underperformance as RP < 0
no_outperformances_1_month = func.no_out_under_performances_ignoring_strings_and_nan(one_month_outperformance_lst)['no_out']
no_outperformances_3_month = func.no_out_under_performances_ignoring_strings_and_nan(three_month_outperformance_lst)['no_out']
no_outperformances_6_month = func.no_out_under_performances_ignoring_strings_and_nan(six_month_outperformance_lst)['no_out']
no_outperformances_lst = [no_outperformances_1_month, no_outperformances_3_month, no_outperformances_6_month]

no_underperformances_1_month = func.no_out_under_performances_ignoring_strings_and_nan(one_month_outperformance_lst)['no_under']
no_underperformances_3_month = func.no_out_under_performances_ignoring_strings_and_nan(three_month_outperformance_lst)['no_under']
no_underperformances_6_month = func.no_out_under_performances_ignoring_strings_and_nan(six_month_outperformance_lst)['no_under']
no_underperformances_lst = [no_underperformances_1_month, no_underperformances_3_month, no_underperformances_6_month]


hit_ratio_1_month = no_outperformances_1_month / (no_outperformances_1_month + no_underperformances_1_month)
hit_ratio_3_month = no_outperformances_3_month / (no_outperformances_3_month + no_underperformances_3_month)
hit_ratio_6_month = no_outperformances_6_month / (no_outperformances_6_month + no_underperformances_6_month)
hit_ratio_lst = [hit_ratio_1_month, hit_ratio_3_month, hit_ratio_6_month]

#Max/Min outperformance
max_1_month_out_performance = func.max_list_ignoring_strings_and_nan(one_month_outperformance_lst)
max_3_month_out_performance = func.max_list_ignoring_strings_and_nan(three_month_outperformance_lst)
max_6_month_out_performance = func.max_list_ignoring_strings_and_nan(six_month_outperformance_lst)
max_out_performance_lst = [max_1_month_out_performance, max_3_month_out_performance, max_6_month_out_performance]

min_1_month_out_performance = func.min_list_ignoring_strings_and_nan(one_month_outperformance_lst)
min_3_month_out_performance = func.min_list_ignoring_strings_and_nan(three_month_outperformance_lst)
min_6_month_out_performance = func.min_list_ignoring_strings_and_nan(six_month_outperformance_lst)
min_out_performance_lst = [min_1_month_out_performance, min_3_month_out_performance, min_6_month_out_performance]


#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>

#Create data frame for averages, median etc.
data = {'Observations': observations_lst, 'Average Performance': avg_out_performance_lst,
'Median Performance': median_out_performance_lst, 'Hit Ratio': hit_ratio_lst,
'Maximum Performance': max_out_performance_lst, 'Minimum Performance': min_out_performance_lst}

#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>
#Create data frame and Excel sheet
Output_file_name = 'OUTPUT/Summary_data.xlsx'

Neste_and_SXXP_summary_df = pd.DataFrame(data, index=pd.Index(['1 month', '3 months', '6 months']))

Neste_and_SXXP_summary_df.to_excel(Output_file_name, sheet_name='Sheet1', startrow = 2, index=True)

#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>
#Formatting Summary Data Spreadsheet
Summary_data_wb = xl.load_workbook(Output_file_name)
Summary_data_ws = Summary_data_wb.active

#Format percentage cells accordingly
percentage_column_letters_lst = ['C', 'D', 'E', 'F', 'G']
for letter in percentage_column_letters_lst:
    for i in range(4, 6+1, 1): #4-7
        Summary_data_ws[letter + str(i)].number_format = '0.00%'

Summary_data_wb.save(Output_file_name)

# print(Neste_and_SXXP_summary_df)
