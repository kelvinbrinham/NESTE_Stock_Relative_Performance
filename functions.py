'''
Here i store some useful functions
'''

import pandas as pd
import numpy as np
import openpyxl as xl
import datetime
from datetime import datetime as dt
import xlsxwriter as xlw
import matplotlib.pyplot as plt
import dateutil.relativedelta as reldel
import statistics as stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill


#Mean of a list ignoring the zero values
def list_mean_ignoring_zeros(list_):
    sum_ = sum(list_)
    length = np.count_nonzero(list_)
    mean = sum_ / length
    return mean

def list_mean_ignoring_strings_and_nan(list_):
    dummy_list = [x for x in list_ if type(x) != str and not np.isnan(x)]
    sum_ = sum(dummy_list)
    length = len(dummy_list)
    mean = sum_ / length
    return mean


def list_median_ignoring_strings_and_nan(list_):
    dummy_list = [x for x in list_ if type(x) != str and not np.isnan(x)]

    return stats.median(dummy_list)


def list_median_ignoring_zeros(list_):
    dummy_list = [x for x in list_ if x != 0]

    return stats.median(dummy_list)


def min_list_ignoring_zeros(list_):
    dummy_list = [x for x in list_ if x != 0]

    return min(dummy_list)


def no_out_under_performances_ignoring_strings_and_nan(list_):
    dummy_list = [x for x in list_ if type(x) != str and not np.isnan(x)]
    no_out = len([x for x in dummy_list if x >= 0])
    no_under = len([x for x in dummy_list if x < 0])
    dict_ = {'no_out': no_out, 'no_under': no_under}

    return dict_



def max_list_ignoring_strings_and_nan(list_):
    dummy_list = [x for x in list_ if type(x) != str and not np.isnan(x)]
    return max(dummy_list)


def min_list_ignoring_strings_and_nan(list_):
    dummy_list = [x for x in list_ if type(x) != str and not np.isnan(x)]
    return min(dummy_list)


def length_ignoring_strings_and_nan(list_):
    dummy_list = [x for x in list_ if type(x) != str and not np.isnan(x)]
    return len(dummy_list)


def colour_yes_cells_yellow(path):
    workbook = xl.load_workbook(path)
    rows_lst = []
    # workbook = Workbook()
    sheet = workbook.active
    yellow = "00FFFF00"
    for rows in sheet.iter_rows(min_row=4, max_row=2634, min_col=1, max_col=10):
        for cell in rows:
            if cell.value == 'Yes':
                rows_lst.append(rows)

    for rows in rows_lst:
        for cell in rows:
            cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
    workbook.save(path)
